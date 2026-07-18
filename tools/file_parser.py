"""
文件解析工具

扩展原 file_reader 的能力，支持 GAIA L1 常见的文件类型：
- PDF：提取文本内容（优先 PyMuPDF，回退 pdfminer.six）
- Excel：提取所有工作表为文本（openpyxl）
- CSV：读取并返回表格内容（内置 csv）
- 图片：返回 base64 编码 + 提示用多模态模型分析
- 纯文本：直接返回内容（兼容原 file_reader）

Executor 在 Planner 规划出 file_parse 步骤时调用本工具。
"""
import os
import base64


def file_parser(args: dict) -> str:
    """
    文件解析工具

    参数:
        args: {
            "path": "文件路径",
            "sheet": "Excel工作表名(可选)",
            "max_chars": "最大返回字符数(默认8000)"
        }

    返回:
        解析后的内容字符串
    """
    path = args.get("path", "")
    max_chars = int(args.get("max_chars", 8000))

    if not path:
        return "错误：缺少 path 参数"

    # 路径安全校验
    safe_base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    real_path = os.path.realpath(path)
    forbidden_prefixes = [
        "/etc", "/var", "/root", "/proc", "/sys", "/dev",
        "C:\\Windows", "C:\\Users", "C:\\Program",
    ]
    for prefix in forbidden_prefixes:
        if real_path.lower().startswith(prefix.lower()):
            return "错误：禁止访问系统敏感路径"
    path_parts = os.path.normpath(path).split(os.sep)
    if any(part.startswith(".") for part in path_parts):
        return "错误：禁止访问隐藏文件"
    if not os.path.exists(path):
        return f"错误：文件不存在 '{path}'"
    if not os.path.isfile(path):
        return f"错误：路径不是文件 '{path}'"
    if os.path.getsize(path) > 20 * 1024 * 1024:
        return f"错误：文件过大 ({os.path.getsize(path)} bytes)，最大支持 20MB"

    ext = os.path.splitext(path)[1].lower()

    try:
        if ext in (".pdf",):
            return _parse_pdf(path, max_chars)
        elif ext in (".xlsx", ".xlsm", ".xls"):
            return _parse_excel(path, args.get("sheet"), max_chars)
        elif ext in (".csv",):
            return _parse_csv(path, max_chars)
        elif ext in (".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"):
            return _parse_image(path, max_chars)
        else:
            # 纯文本回退
            return _parse_text(path, max_chars)
    except Exception as e:
        return f"错误：解析文件失败 [{ext}]: {type(e).__name__}: {str(e)[:200]}"


def _parse_text(path: str, max_chars: int) -> str:
    for enc in ["utf-8", "gbk", "latin-1", "utf-16"]:
        try:
            with open(path, "r", encoding=enc) as f:
                content = f.read()
            return f"[文本文件 {os.path.basename(path)}]\n{content[:max_chars]}"
        except UnicodeDecodeError:
            continue
    return f"错误：无法解码文本文件 '{path}'"


def _parse_pdf(path: str, max_chars: int) -> str:
    text = ""
    # 优先 PyMuPDF
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(path)
        parts = []
        for page in doc:
            parts.append(page.get_text())
        text = "\n".join(parts)
        doc.close()
        return f"[PDF文件 {os.path.basename(path)}，共{doc.page_count if False else len(parts)}页]\n{text[:max_chars]}"
    except ImportError:
        pass
    # 回退 pdfminer.six
    try:
        from pdfminer.high_level import extract_text
        text = extract_text(path)
        return f"[PDF文件 {os.path.basename(path)}]\n{text[:max_chars]}"
    except ImportError:
        return "错误：PDF解析需要安装 PyMuPDF 或 pdfminer.six（pip install pymupdf）"
    except Exception as e:
        return f"错误：PDF解析失败: {str(e)[:200]}"


def _parse_excel(path: str, sheet: str, max_chars: int) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        out = [f"[Excel文件 {os.path.basename(path)}，工作表: {', '.join(wb.sheetnames)}]"]
        sheets = [sheet] if sheet and sheet in wb.sheetnames else wb.sheetnames
        for name in sheets:
            ws = wb[name]
            out.append(f"\n--- 工作表: {name} ---")
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 200:  # 最多200行
                    out.append("...(数据截断，仅显示前200行)")
                    break
                cells = ["" if c is None else str(c) for c in row]
                out.append(" | ".join(cells))
        wb.close()
        return "\n".join(out)[:max_chars]
    except ImportError:
        return "错误：Excel解析需要安装 openpyxl（pip install openpyxl）"
    except Exception as e:
        return f"错误：Excel解析失败: {str(e)[:200]}"


def _parse_csv(path: str, max_chars: int) -> str:
    import csv
    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            rows = []
            for i, row in enumerate(reader):
                if i >= 200:
                    rows.append("...(数据截断，仅显示前200行)")
                    break
                rows.append(" | ".join("" if c is None else str(c) for c in row))
            header = rows[0] if rows else ""
            body = "\n".join(rows)
            return f"[CSV文件 {os.path.basename(path)}]\n列: {header}\n{body[:max_chars]}"
    except Exception as e:
        # 回退纯文本
        return _parse_text(path, max_chars)


def _parse_image(path: str, max_chars: int) -> str:
    # 图片需要多模态模型，这里返回 base64 提示
    try:
        with open(path, "rb") as f:
            b64 = base64.b64encode(f.read()).decode("ascii")
        return (f"[图片文件 {os.path.basename(path)}]\n"
                f"图片已编码为base64（长度 {len(b64)} 字符）。\n"
                f"请将此 base64 传给支持视觉的模型进行分析。\n"
                f"BASE64_START:{b64[:max_chars]}BASE64_END")
    except Exception as e:
        return f"错误：图片读取失败: {str(e)[:200]}"
